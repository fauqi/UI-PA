#from tkinter import *
from numpy import False_, append
from tkinter.constants import S, X
from PIL import ImageTk, Image
from tkinter import OptionMenu, StringVar, messagebox,ttk,Tk,Frame,Label,Button,Entry,PhotoImage,END,Toplevel,NW,CENTER,filedialog
import math
import os
import threading
import time
import subprocess
import serial
from serial import Serial
import struct
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font,colors
import tkinter as tk
from pandas import DataFrame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import ctypes
ctypes.windll.shcore.SetProcessDpiAwareness(1)
countPlot=0
x=0
y=29
Waktu=[x]
Suhu=[y]
data1 = {'Waktu': Waktu,
         'Suhu': Suhu
        }
df1 = DataFrame(data1,columns=['Waktu','Suhu'])



#haruse nng kene
def read_serial():
    global ser
    try:
        ser = serial.Serial('COM7',9600)
        data = ser.readline(5)
        print("USB COM7 Detected")
        print(data)
    except:
        try:
            ser = serial.Serial('COM8',9600)
            data = ser.readline(5)
            print("USB COM8 Detected")
            print(data)
        except:
            try:
                ser = serial.Serial('COM9',9600)
                data = ser.readline(5)
                print("USB COM9 Detected")
                print(data)
            except:
                try:
                    ser = serial.Serial('COM13',9600)
                    data = ser.readline(5)
                    print("USB COM13 Detected")
                    print(data)
                except:
                    pass
                    # print("no USB connected")
read_serial()
windowPage=0

book=Workbook()
sheet=book.active
font_style = Font(bold=True)
sheet['A1']="tanggal"
sheet['B1']="HM"
sheet['C1']="suhu"
sheet['D1']="tegangan"
sheet['E1']="sudut penyalaan"
sheet['F1']="error"
sheet['G1']="derror"
sheet['H1']="out fis"
sheet['I1']="alarm"
sheet['J1']="fan cond"
sheet['K1']="lamp cond"
sheet['L1']="fan state"
sheet['M1']="lamp state"
x=0
fulltext=[0 for x in range(88)]  
flag=0
scaleW=1
scaleH=0.9
count =0
hours=0
count_logging=2
days=0
minutes=0
seconds=0
flag_HM=0
suhu="29"
tegangan=""
sudut_penyalaan=""
error=""
derror=""
out_fuzzy=""
tanggal=""
HM=""
alarm=""
fan_state=""
lamp_state=""
fan_cond=""
lamp_cond=""
log_tanggal=""




root=Tk()
dpi = ctypes.windll.user32.GetDpiForWindow(root.winfo_id())

SCREENWIDTH_unscaled = int(root.winfo_screenwidth())
SCREENHEIGHT_unscaled = int(root.winfo_screenheight())
SCREENWIDTH = int(root.winfo_screenwidth()*scaleW)
SCREENHEIGHT = int(root.winfo_screenheight()*scaleH)
root.state("zoomed")
root.overrideredirect(False)
root.geometry("{0}x{1}+0+0".format(SCREENWIDTH, SCREENHEIGHT))
root.iconbitmap('logo.ico')
def clear(s):
    result=""
    for i in s:
        if i !='[' and i !=']':
            result=result+i
    return(result)

# root.resizable(True,True)
class FullScreenApp(object):
    def __init__(self, master, **kwargs):
        self.master=master
        self.flag=0
        pad=0
        
        # master.bind('<Escape>',self.escape)
        # master.bind('<F>',self.full)
        # master.bind('<f>',self.full)      
    def escape(self,event):
        global SCREENHEIGHT,SCREENWIDTH,scaleH,scaleW
        # if self.flag==0:
        self.master.overrideredirect(False)
        self.flag=1
        scaleH=0.9

            
        Page(root)
        
    def full(self,event):
        global SCREENHEIGHT,SCREENWIDTH,scaleH,scaleW
        self.master.overrideredirect(True)
        self.flag=0
        scaleH=1
        
        Page(root)
        
      
app = FullScreenApp(root)

def pharsing(j):
    global suhu,tegangan,sudut_penyalaan,error,derror,out_fuzzy,log_tanggal,hours,minutes,days,df1,data1,y,x,Waktu,Suhu,count_logging,countPlot
    # data = x.split(",")
    listData=str(j).split(",")
    print(j)
    # listData[0]="b'$fauqi"
    # listData[1]="20"
    # listData[2]="30"
    # listData[3]="40"
    # listData[4]="50"
    # listData[5]="60"
    # listData[6]="70"

    if listData[0]=="b'$fauqi":
        suhu=listData[1]
        if x == 0:
            Suhu = [float(suhu)]
        tegangan=listData[2]
        sudut_penyalaan=listData[3]
        error=listData[4]
        derror=listData[5]
        out_fuzzy=listData[6]
        x=x+1
        y=x*x
        countPlot=countPlot+1
        if countPlot >= 2:
            if x<=22000:
                try:
                    Suhu.append(float(suhu))
                    Waktu.append(x)
                    print(x)
                    data1 = {'Waktu': Waktu,
                    'Suhu': Suhu
                    }
                    df1 = DataFrame(data1,columns=['Waktu','Suhu'])

                    screen.plotting()
                    countPlot=0
                    print("kene")
                except:
                    print("full")
            

        # print("suhu=" + suhu)
        # print("tegangan="+ tegangan)
        # print("sudut penyalaan=" + sudut_penyalaan)
        # print("error=" + error)
        # print("derror=" + derror)
        # print("out fuzzy="+ out_fuzzy)

    else :
        print("pharser failed")
def log_data():
    global suhu,tegangan,sudut_penyalaan,error,derror,out_fuzzy,count_logging,log_tanggal,hours,minutes,days,seconds
    HM=str(days)+":"+str(hours)+":"+str(minutes)+":"+str(seconds)
    sheet.cell(row=count_logging,column=1).value=str(log_tanggal)
    sheet.cell(row=count_logging,column=2).value=str(HM)
    sheet.cell(row=count_logging,column=3).value=str(suhu)
    sheet.cell(row=count_logging,column=4).value=str(tegangan)
    sheet.cell(row=count_logging,column=5).value=str(sudut_penyalaan)
    sheet.cell(row=count_logging,column=6).value=str(error)
    sheet.cell(row=count_logging,column=7).value=str(derror)
    sheet.cell(row=count_logging,column=8).value=str(out_fuzzy)
    count_logging=count_logging+1
    



def date_picker():
    global log_tanggal
    hari=""
    a=dt.date.today()
    b=dt.datetime.now()
    date = b.strftime("%d/%m/%Y %H:%M:%S")
    listData=str(date).split(" ")
    # listData=str(a).split(" ")
    if a.weekday()==0:
        hari = "Senin"
    elif a.weekday()==1:
        hari = "Selasa"
    elif a.weekday()==2:
        hari = "Rabu"
    elif a.weekday()==3:
        hari = "Kamis"
    elif a.weekday()==4:
        hari = "Jumat"
    elif a.weekday()==5:
        hari = "Sabtu"
    elif a.weekday()==6:
        hari = "Minggu"
    tanggal = hari+","+listData[0]
    log_tanggal=listData[0]
    jam=listData[1]+" WIB"
    # print(tanggal)
    # print(jam)
    return tanggal,jam


class Page:
    def __init__(self,master):
        global SCREENHEIGHT,SCREENWIDTH,scaleH,scaleW,df1,dpi
        SCREENWIDTH = int(root.winfo_screenwidth()*scaleW)
        SCREENHEIGHT = int(root.winfo_screenheight()*scaleH)
        master.geometry("{0}x{1}+0+0".format(SCREENWIDTH, SCREENHEIGHT))
  
        
        self.master=master
        self.master.title("UI TUGAS AKHIR FAUQI")
        self.sW=SCREENWIDTH
        self.sH=SCREENHEIGHT
        self.frame=Frame(self.master,bg="RED")
        self.frame2=Frame(self.master,bg="RED")
        self.dpi=dpi
        
        self.row=0
        self.indeks=0
        self.page_init()
        self.showLayar()
        self.master.bind('<Enter>',self.off)
        self.plotting()


    def plotting(self):
        try:
            print("awal plott")
            self.figure1 = plt.Figure(dpi=dpi,frameon=False)
            self.figure1.set_size_inches(self.sW*0.4776/dpi,self.sH*0.51666/dpi)
            self.ax1 = self.figure1.add_subplot(111)
            self.df1 = df1[['Waktu','Suhu']].groupby('Waktu').sum()
            try:
                self.df1.plot(kind='line', legend=True, ax=self.ax1, color='b',marker='o', fontsize=10)
                self.ax1.set_title('Output Respon Suhu')
                bar1 = FigureCanvasTkAgg(self.figure1, self.frame2)
                bar1.get_tk_widget().place(x=0.0755*self.sW,y=0.1814*self.sH)
            except:
                print("mboh")
            print("akhir plot")
        except:
            print("wes entek")
    def off(self,event):
        global proc
        threadPdf.clear()
        self.unloading()

    def page_init(self):


        x=0
        self.clicked= StringVar()
        self.clicked.set("40")
        self.Giflabel = Label(root)
        self.frame.place_forget()
        self.frame2.place_forget()
        self.photo=Image.open("foto/awal.png")
        self.photo = self.photo.resize((self.sW, self.sH), Image.ANTIALIAS)
        self.gambar = ImageTk.PhotoImage(self.photo)
        self.photo2=Image.open("foto/tab2.png")
        self.photo2 = self.photo2.resize((self.sW, self.sH), Image.ANTIALIAS)
        self.gambar2 = ImageTk.PhotoImage(self.photo2)
        self.photo3=Image.open("foto/back.png")
        self.photo3 = self.photo3.resize((int(self.sW*0.032), int(self.sH*0.06)), Image.ANTIALIAS)
        self.backImage = ImageTk.PhotoImage(self.photo3)
        self.photo4=Image.open("foto/alarm_red.png")
        self.photo4 = self.photo4.resize((int(self.sW*0.04166), int(self.sH*0.07407)), Image.ANTIALIAS)
        self.alarm_redImage = ImageTk.PhotoImage(self.photo4)
        self.photo5=Image.open("foto/alarm_green.png")
        self.photo5 = self.photo5.resize((int(self.sW*0.04166), int(self.sH*0.07407)), Image.ANTIALIAS)
        self.alarm_greenImage = ImageTk.PhotoImage(self.photo5)

        self.labelImage=Label(self.frame,height=SCREENHEIGHT,width=SCREENWIDTH,image=self.gambar)
        self.labelImage2=Label(self.frame2,height=SCREENHEIGHT,width=SCREENWIDTH,image=self.gambar2)
        self.alarm_green=Label(self.frame2,image=self.alarm_greenImage,activebackground="#E8EB5B",bg="#E8EB5B",borderwidth=0)
        self.alarm_redn=Label(self.frame2,image=self.alarm_redImage,activebackground="#E8EB5B",bg="#E8EB5B",borderwidth=0)
        self.exitButton = Button(self.frame,command=self.exit,bg="#FE6464",text="EXIT",font='Salsa 25 bold')
        self.exitButton2 = Button(self.frame2,command=self.exit,bg="#FE6464",text="EXIT",font='Helvetica 18 bold')
        self.startButton = Button(self.frame,command=self.start,bg="#42EA27",text="START",font='Salsa 25 bold')
        self.backBtn=Button(self.frame2,image=self.backImage,command=self.back)
        self.loadBtn=Button(self.frame2,command=self.load_data,bg="#9561EB",text="LOAD DATA",font='Helvetica 22 bold')
        self.spBtn=Button(self.frame2,text="set",bg="#9561EB",command=self.sp)
        self.reset_HM=Button(self.frame2,text="Reset",bg="#C4C4C4",command=self.reset_HM,font='Helvetica 12 bold')
        self.HfanBtn=Button(self.frame2,text="H",bg="#C4C4C4",command=self.Hfan,font='Helvetica 12 bold')
        self.OfanBtn=Button(self.frame2,text="O",bg="#C4C4C4",command=self.Ofan,font='Helvetica 12 bold')
        self.AfanBtn=Button(self.frame2,text="A",bg="#C4C4C4",command=self.Afan,font='Helvetica 12 bold')
        self.HlampBtn=Button(self.frame2,text="H",bg="#C4C4C4",command=self.Hlamp,font='Helvetica 12 bold')
        self.OlampBtn=Button(self.frame2,text="O",bg="#C4C4C4",command=self.Olamp,font='Helvetica 12 bold')
        self.AlampBtn=Button(self.frame2,text="A",bg="#C4C4C4",command=self.Alamp,font='Helvetica 12 bold')

        self.labelSuhu=Label(self.frame2)
        self.spMenu=OptionMenu(self.frame2,self.clicked,"40","50","61")
        self.labelDate=Label(self.frame,font='Helvetica 12',bg="#4591EA")
        self.labelTime=Label(self.frame,font='Helvetica 12',bg="#4591EA")
        self.labelDate2=Label(self.frame2,font='Helvetica 12',bg="#4591EA")
        self.labelTime2=Label(self.frame2,font='Helvetica 12',bg="#4591EA")
        self.HM_days=Label(self.frame2,font='Helvetica 12',bg="#ffffff")
        self.HM_hours=Label(self.frame2,font='Helvetica 12',bg="#ffffff")
        self.HM_minutes=Label(self.frame2,font='Helvetica 12',bg="#ffffff")
        self.teganganLabel=Label(self.frame2,font='Helvetica 12',bg="#7BD152")
        self.firingAngleLabel=Label(self.frame2,font='Helvetica 12',bg="#7BD152")
        self.errorLabel=Label(self.frame2,font='Helvetica 12',bg="#7BD152")
        self.derrorLabel=Label(self.frame2,font='Helvetica 12',bg="#7BD152")
        self.outFuzzyLabel=Label(self.frame2,font='Helvetica 12',bg="#7BD152")
    def remove_excel(self):
        global sheet,book,count_logging
        count_logging=count_logging+2
        for n in range(count_logging):
            if n >=2:
                sheet.delete_rows(n)
                print(n)
        count_logging =2
        # book.remove(book['Sheet'])
        # create_workbook()
                

    def load_data(self):
        global book
        file_path=filedialog.asksaveasfile(defaultextension='.xlsx',filetypes=[("xlsx file",".xlsx"),])
        book.save(file_path.name)
        messagebox.showinfo(title="Logging Succes", message="Data Logger sudah tersimpan")
        # print(file_path.name)

    def Hfan(self):
        self.HfanBtn.config(bg="#42EA27")
        self.OfanBtn.config(bg="#C4C4C4")
        self.AfanBtn.config(bg="#C4C4C4")
        ser.write(b"10")
    def Ofan(self):
        self.HfanBtn.config(bg="#C4C4C4")
        self.OfanBtn.config(bg="#42EA27")
        self.AfanBtn.config(bg="#C4C4C4")
        ser.write(b"11")
    def Afan(self):
        self.HfanBtn.config(bg="#C4C4C4")
        self.OfanBtn.config(bg="#C4C4C4")
        self.AfanBtn.config(bg="#42EA27")
        ser.write(b"12")
    def Hlamp(self):
        self.HlampBtn.config(bg="#42EA27")
        self.OlampBtn.config(bg="#C4C4C4")
        self.AlampBtn.config(bg="#C4C4C4")
        ser.write(b"20")
    def Olamp(self):
        self.HlampBtn.config(bg="#C4C4C4")
        self.OlampBtn.config(bg="#42EA27")
        self.AlampBtn.config(bg="#C4C4C4")
        ser.write(b"21")
    def Alamp(self):
        self.HlampBtn.config(bg="#C4C4C4")
        self.OlampBtn.config(bg="#C4C4C4")
        self.AlampBtn.config(bg="#42EA27")   
        ser.write(b"22")
    def reset_HM(self):
        global seconds,minutes,hours,days
        a=messagebox.askyesno(title="reset?",message="Apakah anda yakin ingin mereset Hour Meter?")
        if a == True:
            seconds=0
            minutes=0
            hours=0
            days=0
            self.remove_excel()
            
    def sp(self):
        if self.clicked.get()=="40":
            # a=40
            ser.write("40".encode())
            # ser.write(b"40")
            # ser.write(bytes("40\r", encoding='ascii'))
        elif self.clicked.get()=="50":
            ser.write("50".encode())
            # ser.write(bytes("50\r", encoding='ascii'))
            # ser.write(b"50\r\n")
        elif self.clicked.get()=="61":
            ser.write("61".encode())
            # ser.write(bytes("61\r", encoding='ascii'))
            # ser.write(b"61\r\n")
            
        #ser.write(b"14")
        print(self.clicked.get())
        self.labelSuhu.config(text=self.clicked.get())
        

    def close(self):
        self.frame3.place_forget()
    def unloading(self):
        self.Giflabel.place_forget()
    def showLayar(self):
        global windowPage
        if windowPage==0:
            self.frame.place(x=0,y=0,height=SCREENHEIGHT,width=SCREENWIDTH)
            self.labelImage.place(x=0,y=0,height=SCREENHEIGHT,width=SCREENWIDTH)
            self.exitButton.place(x=self.sW*0.5775 ,y=self.sH*0.7046,width=self.sW*0.1645,height=self.sH*0.0824)
            self.startButton.place(x=self.sW*0.28958 ,y=self.sH*0.7046,width=self.sW*0.1645,height=self.sH*0.0824)
            self.labelDate.place(x=self.sW*0.82,y=self.sH*0.0231,width=self.sW*0.11,height=self.sH*0.0421)
            self.labelTime.place(x=self.sW*0.82,y=self.sH*0.0652,width=self.sW*0.11,height=self.sH*0.0421)
        elif windowPage==1:
            self.start()
    def exit(self):
        a=messagebox.askyesno(title="EXIT?",message="Apakah anda yakin ingin menutup aplikasi?")
        if a == True:
            root.destroy()
        
    def start(self):
        global windowPage,flag_HM
        flag_HM=1
        windowPage=1
        self.frame.place_forget()
        self.frame2.place(x=0,y=0,height=SCREENHEIGHT,width=SCREENWIDTH)
        self.labelImage2.place(x=0,y=0,height=SCREENHEIGHT,width=SCREENWIDTH)
        self.backBtn.place(x=self.sW*0.01,y=self.sH*0.02,width=self.sW*0.032, height=self.sH*0.06)
        self.exitButton2.place(x=0.813*self.sW,y=0.86*self.sH,width=0.1661*self.sW,height=0.0666*self.sH)
        self.spBtn.place(x=self.sW*0.7515,y=self.sH*0.4731,width=self.sW*0.03489, height=self.sH*0.0407)
        self.loadBtn.place(x=0.813*self.sW,y=0.7666*self.sH,width=0.1661*self.sW,height=0.0666*self.sH)
        self.labelSuhu.place(x=0.6552*self.sW,y=0.3296*self.sH,width=0.04583*self.sW,height=0.0546*self.sH)
        self.spMenu.place(x=0.6552*self.sW,y=0.4657*self.sH,width=0.04583*self.sW,height=0.0546*self.sH)
        self.labelDate2.place(x=self.sW*0.82,y=self.sH*0.0231,width=self.sW*0.11,height=self.sH*0.0421)
        self.labelTime2.place(x=self.sW*0.82,y=self.sH*0.0652,width=self.sW*0.11,height=self.sH*0.0421)
        self.HM_days.place(x=self.sW*0.6218,y=self.sH*0.8222,width=self.sW*0.03437,height=self.sH*0.032407)
        self.HM_hours.place(x=self.sW*0.6703,y=self.sH*0.8222,width=self.sW*0.03437,height=self.sH*0.032407)
        self.HM_minutes.place(x=self.sW*0.7171,y=self.sH*0.8222,width=self.sW*0.03437,height=self.sH*0.032407)
        self.reset_HM.place(x=self.sW*0.665,y=self.sH*0.8824,width=self.sW*0.05,height=self.sH*0.0287)
        self.teganganLabel.place(x=self.sW*0.1682,y=self.sH*0.7564,width=self.sW*0.07812,height=self.sH*0.0472)
        self.firingAngleLabel.place(x=self.sW*0.1682,y=self.sH*0.8166,width=self.sW*0.07812,height=self.sH*0.0472)
        self.errorLabel.place(x=self.sW*0.1682,y=self.sH*0.8768,width=self.sW*0.07812,height=self.sH*0.0472)
        self.derrorLabel.place(x=self.sW*0.4312,y=self.sH*0.7564,width=self.sW*0.07812,height=self.sH*0.0472)
        self.outFuzzyLabel.place(x=self.sW*0.4312,y=self.sH*0.8166,width=self.sW*0.07812,height=self.sH*0.0472)
        self.alarm_green.place(x=self.sW*0.8796,y=self.sH*0.35277,width=self.sW*0.04166,height=self.sH*0.07507)
        #HOA
        self.HfanBtn.place(x=self.sW*0.8755,y=self.sH*0.45,width=self.sW*0.01979,height=self.sH*0.0305)
        self.OfanBtn.place(x=self.sW*0.9041,y=self.sH*0.45,width=self.sW*0.01979,height=self.sH*0.0305)
        self.AfanBtn.place(x=self.sW*0.9328,y=self.sH*0.45,width=self.sW*0.01979,height=self.sH*0.0305)
        self.HlampBtn.place(x=self.sW*0.8755,y=self.sH*0.5092,width=self.sW*0.01979,height=self.sH*0.0305)
        self.OlampBtn.place(x=self.sW*0.9041,y=self.sH*0.5092,width=self.sW*0.01979,height=self.sH*0.0305)
        self.AlampBtn.place(x=self.sW*0.9328,y=self.sH*0.5092,width=self.sW*0.01979,height=self.sH*0.0305)
        self.Alamp()
        self.Afan()

        self.plotting()
    def back(self):
        global windowPage,seconds,minutes,days,flag_HM,hours
        windowPage=0
        a=messagebox.askyesno(title="Back?",message="Apakah anda yakin ingin kembali ke halaman awal? saat kembali ke halaman awal data logger akan direset")
        if a == True:
            flag_HM=0
            seconds=0
            minutes=0
            hours=0
            days=0
            self.showLayar()
            self.frame2.place_forget()
            self.frame.place(x=0,y=0,height=SCREENHEIGHT,width=SCREENWIDTH)
            self.master.unbind('<Return>')
            self.remove_excel()




screen = Page(root)

frameCnt = 29
frames = [PhotoImage(file='foto/loading gif.gif',format = 'gif -index %i' %(i)) for i in range(frameCnt)]       
def update(ind):
    frame = frames[ind]
    ind += 1
    if ind == frameCnt:
        ind = 0
    screen.Giflabel.configure(image=frame,bg="WHITE")
    screen.frame.after(100, update, ind)

def loadGif():    
    screen.Giflabel.place(x=SCREENWIDTH_unscaled*0.5,y=SCREENHEIGHT_unscaled*0.5,width = 150,height=150,anchor=CENTER)
    screen.frame.after(0, update, 0)


def kill():
    
    screen.unloading()
def timer2():
    global flag,count,hours,days,minutes,seconds,flag_HM,tegangan,sudut_penyalaan,error,derror,out_fuzzy,suhu,ser,flag_HM,x,y,data1,df1
    while True:
        
        if flag_HM == 1:
            seconds=seconds+1
            if seconds>58:
                minutes=minutes+1
                seconds=0
            if minutes>59:
                hours=hours+1
                minutes=0
            if hours>23:
                days=days+1
                hours=0
            screen.HM_minutes.config(text=minutes)
            screen.HM_hours.config(text=hours)
            screen.HM_days.config(text=days)
        date_picker()
        date=date_picker()[0]
        current_time=date_picker()[1]
        screen.labelDate.config(text=date)
        screen.labelTime.config(text=current_time)
        screen.labelDate2.config(text=date)
        screen.labelTime2.config(text=current_time)
        time.sleep(1)
        if flag_HM == 1:
            log_data()

        # print(seconds)

def timer():
    global flag,count,hours,days,minutes,seconds,tegangan,sudut_penyalaan,error,derror,out_fuzzy,suhu,flag_HM
    while True:
        time.sleep(0.1)
        if flag_HM == 1:
            # pharsing("1,2,3,4,5,6,63,7,75,45,54")#dummy
            try:
                data = ser.readline(100)
                # print(data)
                screen.labelSuhu.config(text=suhu)
                pharsing(data)
                screen.teganganLabel.config(text=tegangan)
                screen.firingAngleLabel.config(text=sudut_penyalaan)
                screen.errorLabel.config(text=error)
                screen.derrorLabel.config(text=derror)
                screen.outFuzzyLabel.config(text=out_fuzzy)
            except:
                print("recieve gagal")
                read_serial()
                # messagebox.showerror(title="recieve gagal!",message="no USB Detected")

 
        

            

            

threadPdf=threading.Event()
t1= threading.Thread(target=timer)
t1.start()
t2= threading.Thread(target=timer2)
t2.start()
threadPdf.set()

root.mainloop()